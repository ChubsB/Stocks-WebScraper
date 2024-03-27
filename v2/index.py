from datetime import datetime
import pandas as pd
import requests
import json
from openpyxl import load_workbook

print("Queuing Update...")
file_path = "Stocks.xlsx"
wb = load_workbook(file_path)
companyList = pd.read_excel(file_path, sheet_name="CompanyList", header=None)
for company in companyList[0]:
    print(f"Fetching data for {company}")
    api_url = "https://www.investorslounge.com/Default/SendPostRequest"
    payload = {
        "url": "PriceHistory/GetPriceHistoryCompanyWise",
        "data": json.dumps({
            "company": company,
            "sort": "0",
            "DateFrom": datetime.now().strftime("%d %b %Y"),
            "DateTo": datetime.now().strftime("%d %b %Y"), 
            # "DateFrom": '26 Mar 2024', # 26 Mar 2024 (1 ahead of start date)
            # "DateTo": '27 Mar 2024',  # 04 Mar 2024 (date needed)
            "key": "",
        }),
    }
    response = requests.post(api_url, json=payload)
    if len(response.json()) > 0:
        data = response.json()[0]
        ws = wb[company]
        ws.move_range("G2:M2", rows=1)  # Move the old data down
        for col in range(7, 14):
            source_cell = ws.cell(row=3, column=col)
            target_cell = ws.cell(row=2, column=col)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = source_cell._style
        ws["G2"].value = datetime.strptime(data["Date_"], "%Y-%m-%dT%H:%M:%S")
        ws["G2"].number_format = 'dd-mmm-yy'
        ws["H2"].value = float(data["Open"])
        ws["I2"].value = float(data["High"])
        ws["J2"].value = float(data["Low"])
        ws["K2"].value = float(data["Close"])
        ws["L2"].value = int(data["Volume"])
        ws["L2"].number_format = '#,##0'
        ws["M2"].value = float(data["Change"])
        ws["M2"].number_format = "0.00"
        print('Successful')
    else:
        print('Failed')
try:
    wb.save(file_path)
    print("Update complete for all companies")
except Exception as e:
    print(f"An error occurred: {e}")
