from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

#This is our Stocks list; if you want to add new stocks add it's name in this list
stocks_list = ['AKBL', 'ANL', 'ASC', 'ASL', 'ASTL', 'ATRL', 'AVN', 'BAFL', 'BIPL','BOP', 'CSIL', 'BYCO', 'DCL', 'DOL', 'EFERT', 'EPCL', 'FABL', 'FCCL', 'FCSC', 'FFBL', 'FFC', 'FFL', 'FNEL', 'GATM', 'GGL', 'GGGL', 'HASCOL', 'HUMNL', 'ICIBL', 'ICL', 'ISL', 'JSBL', 'JSCL',
               'KAPCO', 'KEL', 'KOSM', 'LOADS', 'LOTCHEM', 'MDTL', 'MLCF', 'NBP', 'NRSL', 'PACE', 'PAEL', 'PIAA', 'PIBTL', 'PIOC', 'POWER', 'PPL', 'PRL', 'PSX', 'PTC', 'SEPCO', 'SILK', 'SNGP', 'SPL', 'SSGC', 'STCL', 'STPL', 'TELE', 'TPL', 'TREET', 'TRG', 'UNITY', 'WAVES', 'WTL']
#stocks_list_test = ['AKBL']
#Days worth of data to be collected
days = int(input("Enter the amount of days required: "))
start = 2 
days += 1

#This is to make sure the browser itself does not show 
options = Options()
options.headless = True
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

#Excel File
#wb = load_workbook('Stocks.xlsx')
wb = load_workbook(filename='Stocks.xlsm', read_only=False, keep_vba=True)
sheets = wb.sheetnames
print(sheets)
ws = wb[sheets[1]]
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def historicalPriceCollector(days, StockList,ws):
    #Iterating through the list of stock quotes 
    sheet_count = 1
    for i in StockList:
        
        if(i == "SEPCO"):
            sheet_count += 1
            if(sheet_count < len(sheets)):
                ws = wb[sheets[sheet_count]]
            driver.implicitly_wait(3)
            continue
        
        url = "http://www.scstrade.com/stockscreening/SS_CompanySnapShotHP.aspx?symbol=" + i
        driver.get(url)
        content = driver.page_source
        soup = BeautifulSoup(content, "lxml")
        rows =ws.max_row

        #Getting required stock data from the pages using real xpaths 
        
        for y in range(start, days+1):
            date = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[1]").text
            openx = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[2]").text
            high = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[3]").text
            low = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[4]").text
            close = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[5]").text
            volume = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y) + "]/td[6]").text
            close_previous = driver.find_element_by_xpath(
                "/html[1]/body[1]/form[1]/div[4]/div[2]/div[2]/div[2]/div[1]/div[1]/div[3]/div[1]/ul[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[3]/div[1]/table[1]/tbody[1]/tr[" + str(y+1) + "]/td[5]").text
            change = round(float(close) - float(close_previous), 2)
            print(i+": ", date, openx, high, low, close, volume, change)
            ws.move_range("G"+str(y)+":M"+str(rows), rows=1, cols=0, translate=True)
            ws["G"+str(y)] = date
            ws["G"+str(y)].font = Font(name='Arial', size=10)
            ws["G"+str(y)].alignment = Alignment(horizontal='right')
            ws["G"+str(y)].border = thin_border
            ws["H"+str(y)] = float(openx)
            ws["H"+str(y)].font = Font(name='Arial', size=10)
            ws["H"+str(y)].alignment = Alignment(horizontal='right')
            ws["H"+str(y)].border = thin_border
            ws["I"+str(y)] = float(high)
            ws["I"+str(y)].font = Font(name='Arial', size=10)
            ws["I"+str(y)].alignment = Alignment(horizontal='right')
            ws["I"+str(y)].border = thin_border
            ws["J"+str(y)] = float(low)
            ws["J"+str(y)].font = Font(name='Arial', size=10)
            ws["J"+str(y)].alignment = Alignment(horizontal='right')
            ws["J"+str(y)].border = thin_border
            ws["K"+str(y)] = float(close)
            ws["K"+str(y)].font = Font(name='Arial', size=10)
            ws["K"+str(y)].alignment = Alignment(horizontal='right')
            ws["K"+str(y)].border = thin_border
            volume = volume.replace(",", "")
            ws["L"+str(y)] = int(volume)
            ws["L"+str(y)].font = Font(name='Arial', size=10)
            ws["L"+str(y)].alignment = Alignment(horizontal='right')
            ws["l"+str(y)].border = thin_border
            ws["M"+str(y)] = float(change)
            ws["M"+str(y)].font = Font(name='Arial', size=10)
            ws["M"+str(y)].alignment = Alignment(horizontal='right')
            ws["M"+str(y)].border = thin_border
        sheet_count += 1
        if(sheet_count < len(sheets)):
            ws = wb[sheets[sheet_count]]
        driver.implicitly_wait(3)
    driver.quit()
    wb.save(filename = 'Stocks.xlsm')
    print("All required historical prices have been fetched")
    x = input("Press Enter to exit")
    
historicalPriceCollector(days, stocks_list,ws)